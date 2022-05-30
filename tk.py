import tkinter
from PIL import ImageTk, Image
from tkinter import Tk, Label, Button,Entry, Frame, END , messagebox ,ttk
from tkinter import * 
import cv2
import os
import imutils
import datetime
import time
import numpy as np
import pandas as pd
import openpyxl
import serial
import datetime
from openpyxl import load_workbook
from openpyxl.chart import Reference,  LineChart
from git import Repo
import PIL.Image
import PIL.ImageTk

# formato = '%c'
# ahora = time.strftime(formato)
ahora=datetime.datetime.now()
ahora1=ahora.strftime("%d/%m/%y")
ahora2=ahora.strftime("%Hh/%Mm/%Ss")

fechaActual= datetime.datetime.now()






def ventanaregis():
    
    ventana.withdraw()
    ventana2 = tkinter.Toplevel()
#     ventana2.geometry('600x400')
    ventana2.attributes('-fullscreen', True)
    ventana2.title('REGISTRO')
    labelregistro = Label(ventana2, image=photo1).place(x=0,y=0,relwidth=1.0,relheight=1.0)

#     labelregistro.pack()

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #     

   
#########################################################################################
#########################################################################################
    entradanombre= tkinter.Entry(ventana2, text= 'Nombre',font=('ALGERIAN 20 bold'))
#     entradanombre.grid(row= 1, column = 1)
    entradanombre.place(x=456 , y=207 , width=652, height=70)

    entradaapellido= tkinter.Entry(ventana2, text= 'Apellido',font=('ALGERIAN 20 bold'))
#     entradaapellido.grid(row= 2, column = 1)
    entradaapellido.place(x=456 , y=291 , width=652, height=70)

    entradacedula= tkinter.Entry(ventana2, text= 'Cédula',font=('ALGERIAN 20 bold'))
#     entradacedula.grid(row= 3, column = 1)
    entradacedula.place(x=456 , y=376 , width=652, height=70)
    
    entradacurso = ttk.Combobox(ventana2, state = 'readonly',font=('ALGERIAN 20 bold'))
    entradacurso['values']=['Seleccione','Inicial 1','Inicial 2','Primero','Segundo','Tercero','Cuarto','Quinto','Sexto','Séptimo','Octavo','Noveno','Décimo','Primero de Bachillerato','Segundo de Bachillerato','Tercero de Bachillerato','Básica Acelerado','Docente','Otros']
    entradacurso.current(0)
#     entradacurso.grid(row= 4, column = 1 )
    entradacurso.place(x=456 , y=462 , width=652, height=70)

    
#     entradacurso= tkinter.Entry(ventana2, text= 'Curso')
#     entradacurso.grid(row= 4, column = 1)
    

    def entrenar():
        
        dataPath = '/home/pi/DispositivoFinalIoMT/Data' #Cambia a la ruta donde hayas almacenado Data
        peopleList = os.listdir(dataPath)
        print('Lista de personas: ', peopleList)

        labels = []
        facesData = []
        label = 0

        for nameDir in peopleList:
            personPath = dataPath + '/' + nameDir
            print('Leyendo las imágenes')

            for fileName in os.listdir(personPath):
                print('Rostros: ', nameDir + '/' + fileName)
                labels.append(label)
                facesData.append(cv2.imread(personPath+'/'+fileName,0))

            label = label + 1

        face_recognizer = cv2.face.EigenFaceRecognizer_create()

        print("Entrenando...")
        face_recognizer.train(facesData, np.array(labels))


        face_recognizer.write('modeloEigenFacepruebainterface.xml')

        print("Modelo almacenado...")

    def mensajeregistro():
        messagebox.showinfo('Registro','Registro Realizado Correctamente')
    def mensajeerror():
        messagebox.showinfo('Error','El usuario que desea Registrar ya existe en la base de datos')
    
    def atras():
        ventana2.withdraw()
        ventana.deiconify()
        
    def capturarrostro():
        nombre = entradanombre.get()
        apellido = entradaapellido.get()
        cedula = entradacedula.get()
        curso = entradacurso.get()
        
        if(entradacurso.get()== 'Seleccione'):
            curso = 'Otros'

        personName = nombre + ' ' + apellido
        dataPath = '/home/pi/DispositivoFinalIoMT/Data' #Cambia a la ruta donde hayas almacenado Data
        personPath = dataPath + '/' + personName
        
        
        
        

        if (os.path.isfile('/home/pi/DispositivoFinalIoMT/Todos los Datos/'+ personName +'.csv')):
            mensajeerror()
        
        else:
            os.path.exists(personPath)
            print('Carpeta creada: ',personPath)
            os.makedirs(personPath)
            
            
            cap = cv2.VideoCapture(0,cv2.CAP_V4L)
            

            faceClassif = cv2.CascadeClassifier(cv2.data.haarcascades+'haarcascade_frontalface_default.xml')
            count = 0

            while True:

                ret, frame = cap.read()
                if ret == False: break
                frame =  imutils.resize(frame, width=640)
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                auxFrame = frame.copy()

                faces = faceClassif.detectMultiScale(gray,1.3,5)

                for (x,y,w,h) in faces:
                    cv2.rectangle(frame, (x,y),(x+w,y+h),(0,255,0),2)
                    rostro = auxFrame[y:y+h,x:x+w]
                    rostro = cv2.resize(rostro,(150,150),interpolation=cv2.INTER_CUBIC)
                    cv2.imwrite(personPath + '/rostro_{}.jpg'.format(count),rostro)
                    count = count + 1
                cv2.imshow('frame',frame)

                k =  cv2.waitKey(1)
                if k == 27 or count >= 20:
                
                    break

            cap.release()
            cv2.destroyAllWindows()
            
            entrenar()
#########################################################################################            
            archivo = open ('/home/pi/DispositivoFinalIoMT/Todos los Datos/'+ personName +'.csv','a')
            archivo.write('Nombre '+ ','+ personName + '\n' + 'Cedula' + ',' + cedula +'\n'+ 'Curso' + ',' + curso +'\n')
            archivo.close()
            
#########################################################################################
            wb = openpyxl.Workbook()
            ws=wb.active
            temp=(["Nombre", personName])
            temp2=(["Cédula", cedula])
            temp3=(["Curso", curso])
            temp5=(["Fecha","Hora", "Temperatura","Peso","Altura","IMC","O2Sat"])
            ws.append(temp)
            ws.append(temp2)
            ws.append(temp3)
            ws.append(temp5)
            wb.save('/home/pi/GitHub-IoMT/IoMT-BD/Datos Almacenados/'+curso+'/'+personName+".xlsx")
     
            mensajeregistro()
            
            entradanombre.delete(0,END) 
            entradaapellido.delete(0,END) 
            entradacedula.delete(0,END) 
            entradacurso.delete(0,END) 
        


    botonaceptar = tkinter.Button(ventana2, text = 'Aceptar', command  = capturarrostro, padx=74, pady=18,font=('ALGERIAN 20 bold') )
    botonaceptar.place(x=289, y=606)
    #botonregistro.pack()  // aparece el boton en la ventana
    #botonaceptar.grid(row =6, column=1)
    
    botonregresar = tkinter.Button(ventana2, text = 'Regresar', command  = atras , padx=66, pady=18, font=('ALGERIAN 20 bold'))
    botonregresar.place(x=722, y=606)
    #botonregistro.pack()  // aparece el boton en la ventana
    #botonregresar.grid(row =6 , column=2)
    

def actualizargithub():
    
    repo = Repo('/home/pi/GitHub-IoMT/IoMT-BD')  # if repo is CWD just do '.'
    repo.index.add(['/home/pi/GitHub-IoMT/IoMT-BD/Datos Almacenados'])
    repo.index.commit('actualizando')
    origin = repo.remote('origin')
    origin.push()
    
def reconocimientof():
  

    dataPath = '/home/pi/DispositivoFinalIoMT/Data' #Cambia a la ruta donde hayas almacenado Data
    imagePaths = os.listdir(dataPath)
    print('imagePaths=',imagePaths)

    face_recognizer = cv2.face.EigenFaceRecognizer_create()

    face_recognizer.read('modeloEigenFacepruebainterface.xml')


    cap = cv2.VideoCapture(0,cv2.CAP_V4L)


    faceClassif = cv2.CascadeClassifier(cv2.data.haarcascades+'haarcascade_frontalface_default.xml')
    
    count1 = 0
    
    while True:
        ret,frame = cap.read()
        if ret == False: break
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        auxFrame = gray.copy()

        faces = faceClassif.detectMultiScale(gray,1.3,5)
        

        for (x,y,w,h) in faces:
            rostro = auxFrame[y:y+h,x:x+w]
            rostro = cv2.resize(rostro,(150,150),interpolation= cv2.INTER_CUBIC)
            result = face_recognizer.predict(rostro)

            cv2.putText(frame,'{}'.format(result),(x,y-5),1,1.3,(255,255,0),1,cv2.LINE_AA)

            if result[1] < 4500:
                cv2.putText(frame,'{}'.format(imagePaths[result[0]]),(x,y-25),2,1.1,(0,255,0),1,cv2.LINE_AA)
                cv2.rectangle(frame, (x,y),(x+w,y+h),(0,255,0),2)
                nombrefinal = '{}'.format(imagePaths[result[0]])
                nombreparallenar = str(nombrefinal)
                count1 = count1 + 1

            else:
                cv2.putText(frame,'Desconocido',(x,y-20),2,0.8,(0,0,255),1,cv2.LINE_AA)
                cv2.rectangle(frame, (x,y),(x+w,y+h),(0,0,255),2)
                
        cv2.imshow('frame',frame)
        
        print(count1)
        
        k =  cv2.waitKey(1)
        if k == 27 or count1 >= 10:
        
            break
        

    
    cap.release()
    cv2.destroyAllWindows()
       

    


 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
 
    datos=pd.read_csv('/home/pi/DispositivoFinalIoMT/Todos los Datos/'+nombreparallenar+'.csv', header= None)
    df=pd.DataFrame(datos)
    a = str(df.iat[2,1])
    print(a)
    
    serialArduino = serial.Serial('/dev/ttyACM0',9600)
    cont1=0
    cont2=0
    while True :
        cad=serialArduino.readline().decode('ascii')
        if(cont1>=10):

            datoss=cad.splitlines()
            d0=str(datoss[0])
            d1=d0.replace("b","")
            d2=d1.replace("'","")
            d3=d2.split(",")
            dato1=str(d3[0])
            dato2=str(d3[1])
            dato3=str(d3[2])
            dato4=str(d3[3])
            dato5=str(d3[4])
            print(dato1,dato2,dato3,dato4,dato5)
            print("--------------------")
            cont1=cont1+1
#             
#             if(dato3<=0):
#                 dato3=1
            
        if(cont2==20):
            dato6=float(dato2)/(float(dato3)*float(dato3))
            dato7=round(dato6,2)
            dato7=2
#             archivo = open ('/home/pi/Desktop/Datos Almacenados/'+a+'/'+nombreparallenar+'.xlsx','a')
#             archivo.write(ahora +  ','+ dato1 +  ','+ dato2 +  ','+ dato3 +  ','+ str(dato7) +  ','+ dato5 +  '\n')
#             archivo.close()
            wb2=openpyxl.load_workbook('/home/pi/GitHub-IoMT/IoMT-BD/Datos Almacenados/'+a+'/'+nombreparallenar+'.xlsx')
            ws=wb2.active
            temp=([ahora1, ahora2, dato1,dato2,dato3,str(dato7),dato5])
            ws.append(temp)
            wb2.save('/home/pi/GitHub-IoMT/IoMT-BD/Datos Almacenados/'+a+'/'+nombreparallenar+'.xlsx')
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

            serialArduino.close()
            cont1=0
            cont2=0
            
            actualizargithub()
            
            break

        cont1=cont1+1
        cont2=cont2+1
        print(cont1,cont2)
        time.sleep(0.5)
    


 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #    


    

    
    messagebox.showinfo('Datos','Datos Almacenados Correctamente')
#     ventana3.withdraw()
#     botonsalir = tkinter.Button(ventana, text = 'Autentificacion', command = reconocimientof)
    ventana.deiconify() 




ventana = tkinter.Tk()
# ventana.geometry('600x600')  #TAma;o de la ventana principal
ventana.title('Ventana Principal')
ventana.attributes('-fullscreen', True)

im = PIL.Image.open("/home/pi/DispositivoFinalIoMT/fondoprincipal.png")
im1 = PIL.Image.open("/home/pi/DispositivoFinalIoMT/fondoregistro.png")
photo = PIL.ImageTk.PhotoImage(im)
photo1 = PIL.ImageTk.PhotoImage(im1)
label = Label(ventana, image=photo).place(x=0,y=0,relwidth=1.0,relheight=1.0)


actualizargithub()

# etiqueta =  tkinter.Label(ventana, text= 'Bienvenido', bg = 'gray',  height=5 , width=40)  #Etiqueta de Entrada 
# #etiqueta.grid(row = 0, column = 4) # Aparece la Linea Gris en toda la parte superior
# etiqueta.place(x=450 , y=20)
botonregistro = tkinter.Button(ventana, text = 'Registro', command  = ventanaregis , padx=74, pady=56,font=('ALGERIAN 20 bold'))
#botonregistro.pack()  // aparece el boton en la ventana
#botonregistro.grid(row =2, column=2)
botonregistro.place(x=203 , y=390)

botonautentificacion = tkinter.Button(ventana, text = 'Autentificación', command = reconocimientof,padx=25, pady=56,font=('ALGERIAN 20 bold'))
#botonautentificacion.pack()
#botonautentificacion.grid(row =1, column=2)
botonautentificacion.place(x=790, y=390)





serialArduino = serial.Serial('/dev/ttyACM0',9600)


ventana.mainloop()